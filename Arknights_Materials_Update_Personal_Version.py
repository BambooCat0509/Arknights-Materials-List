import sys
import argparse
import xlwings as xw
import traceback as tb
import openpyxl as opxl
from datetime import datetime
from selenium import webdriver
from itertools import zip_longest
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support import expected_conditions as EC

def str_to_bool(value):
    BOOLEAN_MAPPING = {
        "TRUE": True, "True": True, "true": True, "T": True, "t": True, "1": True,
        "FALSE": False, "False": False, "false": False, "F": False, "f": False, "0": False
    }
    return BOOLEAN_MAPPING.get(value, True)
def Scroll(driver, direction):
    scroll_height = driver.execute_script("return document.body.scrollHeight")
    scroll_width = driver.execute_script("return document.body.scrollWidth")
    while True:
        if direction == "up":
            driver.execute_script("window.scrollTo(0, 0);")
        elif direction == "down":
            driver.execute_script("window.scrollTo(document.body.scrollWidth, document.body.scrollHeight);")
        else:
            line_number = tb.format_exc().split(",")[1].strip().split()[-1]
            print(f"Error: Scroll() 參數無效\nLine: {line_number}, Error: {e}")
            sys.exit(0)
        new_scroll_width = driver.execute_script("return document.body.scrollWidth")
        new_scroll_height = driver.execute_script("return document.body.scrollHeight")
        if new_scroll_height == scroll_height and new_scroll_width == scroll_width:
            break
        scroll_height = new_scroll_height
        scroll_width = new_scroll_width

if __name__ == '__main__':
    # 接收輸入的參數
    parser = argparse.ArgumentParser(usage=argparse.SUPPRESS, add_help=False)
    parser.add_argument("--Help", "-H", "-h", action="help", help="show this help message and exit")
    parser.add_argument("--ChromeDriverPath", "--Driver", "-D", "-d", metavar="", default="C:\Program Files\Google\Chrome\Application\chromedriver-win64\chromedriver.exe", help="Chrome Driver 的絕對路徑")
    parser.add_argument("--FilePath", "--File", "-F", "-f", metavar="", default="D:\Films_Edit\明日方舟\Arknights\明日方舟 素材一覽.xlsx", help="\"明日方舟 素材一覽.xlsx\" 的絕對路徑")
    parser.add_argument("--NoCount", "--NC", "-N", "-n", nargs="*", metavar="", help="不想列入計算的關卡代號 / 活動名稱 (簡體字或英文，多個以空格區分，ex: -n 长夜临光 CW)")
    parser.add_argument("--Minimun", "--Min", "-M", "-m", metavar="", type=str_to_bool, default=True, help="是否獲取素材單件最低期望理智 (True: 1, False: 0)")
    parser.add_argument("--Comprehensive", "--Com", "-C", "-c", metavar="", type=str_to_bool, default=True, help="是否計算綜合素材最高效率關卡 (True: 1, False: 0)")
    if not parser.parse_args().Minimun and not parser.parse_args().Comprehensive:
        print("--Comprehensive和--Minimun間，至少需一項操作為True！")
        sys.exit(0)
    CHROME_DRIVER_PATH = rf"{parser.parse_args().ChromeDriverPath}"
    FILE_PATH = rf"{parser.parse_args().FilePath}"
    NoCount = parser.parse_args().NoCount
    if not NoCount:
        NoCount = []
    NoCount.append("补给")
    NoCount.append("岁过华灯")
    NoCount.append("奖励扭蛋机")
    
    # 目標網址
    URL = "https://penguin-stats.io/result/item"
    # 網站嘗試重連次數上限
    RECONNECTION_LIMIT = 3
    
    # 檢查檔案是否關閉
    try:
        workbook = opxl.load_workbook(FILE_PATH, data_only = False)
        workbook.save(FILE_PATH)
        workbook.close()
    except KeyboardInterrupt:
        print("Keyboard interruptted!")
        sys.exit(0)
    except Exception as e:
        if "Permission denied" in "{}".format(e):
            print(f"Error: 請先關閉檔案，再繼續執行 \"{FILE_PATH}\"")
        else:
            line_number = tb.format_exc().split(",")[1].strip().split()[-1]
            print(f"Error: 檢查檔案是否關閉\nLine: {line_number}, Error: {e}")
        sys.exit(0)
    
    print(f"\n\n##### 程式運行結束前，請勿開啟檔案\"{FILE_PATH}\" #####\n")
    # 打印開始執行時間
    print(f"Start time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    # 設定瀏覽器
    print("Options setting")
    options = Options()
    options.add_argument("--headless")  # 啟用無頭模式
    options.add_argument("--disable-gpu")  # 禁用 GPU 加速（Windows 可能需要）
    options.add_argument("--log-level=3")  # 設定低級日誌輸出
    options.add_argument("--window-size=1920,1080")  # 設定視窗大小，防止某些網站依賴視窗尺寸渲染內容
    options.add_argument("--disable-gpu-compositing")  # 禁用 GPU 組合
    options.add_argument("--disable-software-rasterizer")  # 禁用軟體光柵化
    options.add_argument("user-agent=Mozilla/5.0 (Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36")
    driver = webdriver.Chrome(service=Service(CHROME_DRIVER_PATH), options=options)
    print("\033[F\033[K" * 4)
    
    # 資料爬蟲
    try:
        # 所有素材頁面
        print("Establishing connection ...")
        driver.get(URL)
        WebDriverWait(driver, timeout=2).until(lambda driver: driver.current_url == URL)
        print("\033[F\033[KConnection established!")
        
        reconnection = 0
        exp = 999
        ski = 999
        car = 999
        materials_all, materials_w, materials_g, materials_b, materials_p, materials_e = [], [], [], [], [], []
        materials_name, materials_min, materials_ratio = [], [], []
        types_factors = {
            "經驗": {"黃": 1,"紫": 2,"藍": 5,"綠": 10},
            "書": {"藍": 1,"綠": 3,"白": 9},
            "碳": {"大": 1,"中": 3,"小": 9},
            "金": {"赤": 72 / 180}
        }
        variables = {
            "經驗": exp,
            "金": exp,
            "書": ski,
            "碳": car
        }
        # 取得並記錄素材單件最低期望理智
        if parser.parse_args().Minimun:
            # 儲存素材名稱
            var = "//div[contains(@class, 'ml-2 my-2') and (contains(text(), '作战记录') or contains(text(), '材料'))]"\
                    "/following-sibling::div[contains(@class, 'd-flex flex-wrap justify-start')]//figure[@alt]"
            elements = driver.find_elements(By.XPATH, var)
            for element in elements:
                materials_name.append(element.get_attribute('alt').strip())
            
            # 素材單件最低期望理智
            print("正在獲取 \"素材單件最低期望理智\" ...\n")
            for index_name, material_name in enumerate(materials_name):
                print(f"\033[F\033[K\t正在獲取素材資訊({index_name+1:0>2d}/{len(materials_name)})：{material_name} ...")
                while reconnection <= RECONNECTION_LIMIT:
                    try:
                        # 跳轉指定素材頁面
                        WebDriverWait(driver, timeout=2).until(lambda driver: driver.current_url == URL)
                        var = f'[alt="{material_name}"]'
                        driver.execute_script("arguments[0].scrollIntoView(true); arguments[0].click();", WebDriverWait(driver, timeout=2, poll_frequency=0.1).until(EC.element_to_be_clickable((By.CSS_SELECTOR, var))))
                        WebDriverWait(driver, timeout=2).until(lambda driver: driver.current_url != URL)
                        material_URL = driver.current_url
                        # 選擇數據過濾方式
                        try:
                            Scroll(driver, "up")
                            var = "//div[contains(@class, 'v-expansion-panel') and @aria-expanded]"
                            element = WebDriverWait(driver, timeout=2, poll_frequency=0.1).until(EC.visibility_of_element_located((By.XPATH, var)))
                            if element.get_attribute("aria-expanded") == "false":
                                element.click()
                                try:
                                    Scroll(driver, "up")
                                    var = ".v-input--selection-controls__input"
                                    element = WebDriverWait(driver, timeout=2, poll_frequency=0.1).until(EC.visibility_of_all_elements_located((By.CSS_SELECTOR, var)))[2]
                                    if element.find_element(By.CSS_SELECTOR, "input").get_attribute("aria-checked") == "false":
                                        element.click()
                                except TimeoutException:
                                    print(f"\tTimeout: 僅正在開放")
                                    sys.exit(0)
                                except Exception as e:
                                    if "ERR_INTERNET_DISCONNECTED" in "{}".format(e):
                                        print(f"Error: 網路連線失敗")
                                    elif "target window already closed" in "{}".format(e):
                                        print(f"Error: 瀏覽器意外關閉")
                                    else:
                                        line_number = tb.format_exc().split(",")[1].strip().split()[-1]
                                        print(f"\tError: 僅正在開放\n\tLine: {line_number}, Error: {e}")
                                    sys.exit(0)
                        except TimeoutException:
                            pass
                        except Exception as e:
                            if "ERR_INTERNET_DISCONNECTED" in "{}".format(e):
                                print(f"Error: 網路連線失敗")
                            elif "target window already closed" in "{}".format(e):
                                print(f"Error: 瀏覽器意外關閉")
                            else:
                                line_number = tb.format_exc().split(",")[1].strip().split()[-1]
                                print(f"\tError: 選擇數據過濾方式\n\tLine: {line_number}, Error: {e}")
                            sys.exit(0)
                        # 選擇數據排列依據
                        try:
                            Scroll(driver, "up")
                            var = "//th[contains(@aria-label, '单件期望理智')]"
                            element = WebDriverWait(driver, timeout=2, poll_frequency=0.1).until(EC.visibility_of_element_located((By.XPATH, var)))
                            if element.get_attribute("aria-sort") != "ascending":
                                element.click()
                        except TimeoutException:
                            print(f"\tTimeout: 選擇數據排列依據")
                            sys.exit(0)
                        except Exception as e:
                            if "ERR_INTERNET_DISCONNECTED" in "{}".format(e):
                                print(f"Error: 網路連線失敗")
                            elif "target window already closed" in "{}".format(e):
                                print(f"Error: 瀏覽器意外關閉")
                            else:
                                line_number = tb.format_exc().split(",")[1].strip().split()[-1]
                                print(f"\tError: 選擇數據排列依據\n\tLine: {line_number}, Error: {e}")
                            sys.exit(0)
                        # 儲存素材最低期望理智
                        var = "span.d-flex.flex-column.ml-2"
                        elements = WebDriverWait(driver, timeout=2, poll_frequency=0.1).until(EC.visibility_of_all_elements_located((By.CSS_SELECTOR, var)))
                        for index_element, element in enumerate(elements):
                            skip = False
                            for str in NoCount:
                                if str in element.text:
                                    skip = True
                                    break
                            if skip:
                                continue
                            stage_min = element.text
                            var = f"//tr[{index_element+1}]/td[@class='d-flex align-center justify-start fill-height px-2 font-weight-bold monospace']/following-sibling::td[@class='px-2 font-weight-bold monospace'][1]"
                            consume_min = WebDriverWait(driver, timeout=2, poll_frequency=0.1).until(EC.visibility_of_element_located((By.XPATH, var))).text
                            break
                        materials_min.append(f"{material_name}\t{stage_min}\t{consume_min}".replace(" ", "").split())
                        # 返回所有素材
                        driver.back()
                        reconnection = 0
                        break
                    except TimeoutException as e:
                        if reconnection < RECONNECTION_LIMIT:
                            reconnection += 1
                            print(f"\tReconneting ({reconnection}/{RECONNECTION_LIMIT}) ...\n")
                            driver.get(URL)
                            continue
                        else:
                            print(f"\tTimeout: {material_name}")
                            sys.exit(0)
                    except Exception as e:
                        if "target frame detached" in "{}".format(e) and reconnection < RECONNECTION_LIMIT:
                            reconnection += 1
                            print(f"\tReconneting ({reconnection}/{RECONNECTION_LIMIT}) ...\n")
                            driver.get(URL)
                            continue
                        else:
                            if "ERR_INTERNET_DISCONNECTED" in "{}".format(e):
                                print(f"Error: 網路連線失敗")
                            elif "target window already closed" in "{}".format(e):
                                print(f"Error: 瀏覽器意外關閉")
                            else:
                                line_number = tb.format_exc().split(",")[1].strip().split()[-1]
                                print(f"\tError: {material_name}\n\tLine: {line_number}, Error: {e}")
                            sys.exit(0)
            
            # 素材單件最低期望理智更新
            print("\033[F\033[K" * 2 + "\"素材單件最低期望理智\" 更新中 ...")
            try:
                app = xw.App(visible=False)
                workbook = app.books.open(FILE_PATH)
                workbook.app.calculate()
                workbook.save()
                workbook.close()
                app.quit()
                workbook = opxl.load_workbook(FILE_PATH, data_only = True)
                sheet = "素材一覽"
                worksheet = workbook[sheet]
                # 取得資料對應表
                materials = [(4, materials_w), (7, materials_g), (10, materials_b), (19, materials_e)]
                for min_col, material in materials:
                    for row in worksheet.iter_rows(min_col=min_col, min_row=2, values_only=True):
                        if row[0] is not None:
                            material.append([row[0], row[1]])
                workbook.close()
                
                # 資料比對、合併
                for data_low in materials_min:
                    for (index_w, name_w), (index_g, name_g), (index_b, name_b), (index_e, name_e) in zip_longest(enumerate(materials_w), enumerate(materials_g), enumerate(materials_b), enumerate(materials_e), fillvalue=(None, None)):
                        if name_w is not None and data_low[0] == name_w[0]:
                            materials_w[index_w] = [float(data_low[3]), name_w[1], data_low[2], data_low[1]]
                        if name_g is not None and data_low[0] == name_g[0]:
                            materials_g[index_g] = [float(data_low[3]), name_g[1], data_low[2], data_low[1]]
                        if name_b is not None and data_low[0] == name_b[0]:
                            materials_b[index_b] = [float(data_low[3]), name_b[1], data_low[2], data_low[1]]
                        if name_e is not None and data_low[0] == name_e[0]:
                            materials_e[index_e] = [float(data_low[3]), name_e[1]]
                for data_w in materials_w:
                    for data_g in materials_g:
                        if data_w[1][1:] == data_g[1][1:] and float(data_w[0]) * 3 < float(data_g[0]):
                            data_g[0] = float(data_w[0]) * 3
                            data_g[2] = data_w[2]
                            data_g[3] = data_w[3]
                            break
                for data_g in materials_g:
                    for data_b in materials_b:
                        if data_g[1][1:] == data_b[1]:
                            if "石頭" not in data_b[1] and float(data_g[0]) * 4 < float(data_b[0]):
                                data_b[0] = float(data_g[0]) * 4
                            elif "石頭" in data_b[1] and float(data_g[0]) * 5 < float(data_b[0]):
                                data_b[0] = float(data_g[0]) * 5
                            data_b[2] = data_g[2]
                            data_b[3] = data_g[3]
                            break
                for data_e in materials_e:
                    factor = types_factors.get(data_e[1][1:], {}).get(data_e[1][0], 1000)
                    if data_e[1] == "赤金":
                        factor = types_factors.get("經驗").get("紫") * types_factors.get("金").get("赤")
                    if data_e[1][1:] in variables:
                        var_name = data_e[1][1:]
                        current_value = variables[var_name]
                        if current_value < float(data_e[0]) * factor:
                            data_e[0] = current_value / factor
                        else:
                            variables[var_name] = float(data_e[0]) * factor
                
                # 寫入數據
                app = xw.App(visible=False)
                workbook = app.books.open(FILE_PATH)
                workbook.app.calculate()
                workbook.save()
                workbook.close()
                app.quit()
                workbook = opxl.load_workbook(FILE_PATH, data_only = False)
                sheet = "素材一覽"
                worksheet = workbook[sheet]
                for index_row, row in enumerate(worksheet.iter_rows(min_col=20, min_row=2, max_row = 4), start = 2):
                    for data_e in materials_e:
                        if data_e[1] == row[0].value:
                            worksheet.cell(row=index_row, column=row[0].column+1, value=float(data_e[0]))
                            break
                sheet = "打素材"
                worksheet = workbook[sheet]
                worksheet.cell(row=1, column=1, value=datetime.now().strftime("%Y-%m-%d"))
                for index_row, row in enumerate(worksheet.iter_rows(min_col=1, min_row=2), start = 2):
                    for data_b in materials_b:
                        if data_b[1] == row[0].value:
                            worksheet.cell(row=index_row, column=row[0].column+1, value=float(data_b[0]))
                            if "标准" in data_b[3]:
                                worksheet.cell(row=index_row, column=row[0].column+2, value=f"{data_b[2]}(標準)")
                            elif "磨难" in data_b[3]:
                                worksheet.cell(row=index_row, column=row[0].column+2, value=f"{data_b[2]}(磨難)")
                            else:
                                worksheet.cell(row=index_row, column=row[0].column+2, value=data_b[2])
                            break
                workbook.save(FILE_PATH)
                workbook.close()
            except Exception as e:
                if "Permission denied" in "{}".format(e):
                    print(f"Error: \"素材單件最低期望理智\" 更新失敗，請先關閉檔案 \"{FILE_PATH}\"")
                else:
                    line_number = tb.format_exc().split(",")[1].strip().split()[-1]
                    print(f"Error: \"素材單件最低期望理智\" 更新\nLine: {line_number}, Error: {e}")
                sys.exit(0)
            print("\033[F\033[K\"素材單件最低期望理智\" 更新完成")
        
        # 取得並記錄綜合素材最高效率關卡
        if parser.parse_args().Comprehensive:
            # 取得紫材&全素材資料對應表
            app = xw.App(visible=False)
            workbook = app.books.open(FILE_PATH)
            workbook.app.calculate()
            workbook.save()
            workbook.close()
            app.quit()
            workbook = opxl.load_workbook(FILE_PATH, data_only = True)
            sheet = "素材一覽"
            worksheet = workbook[sheet]
            for index_row, row in enumerate(worksheet.iter_rows(min_col=13, min_row=2, values_only=True), start=2):
                if row[0] is not None:
                    materials_p.append([row[0], row[1]])
            for _, row in enumerate(worksheet.iter_rows(min_col=1, min_row=2, values_only=True), start=2):
                if row[0] is not None:
                    materials_all.append([row[0], row[1], row[2]])
            workbook.close()
            
            # 綜合素材最高效率關卡
            print("正在計算 \"綜合素材最高效率關卡\" ...\n\n\n")
            for index_m, material_p in enumerate(materials_p):
                while reconnection <= RECONNECTION_LIMIT:
                    try:
                        # 所有素材頁面
                        WebDriverWait(driver, timeout=2).until(lambda driver: driver.current_url == URL)
                        var = f'[alt="{material_p[0]}"]'
                        driver.execute_script("arguments[0].scrollIntoView(true); arguments[0].click();", WebDriverWait(driver, timeout=2, poll_frequency=0.1).until(EC.element_to_be_clickable((By.CSS_SELECTOR, var))))
                        reconnection = 0
                        break
                    except TimeoutException as e:
                        if reconnection < RECONNECTION_LIMIT:
                            reconnection += 1
                            print(f"\tReconneting ({reconnection}/{RECONNECTION_LIMIT}) ...\n")
                            driver.get(URL)
                            continue
                        else:
                            print(f"\tTimeout: 企鵝物流 - 素材選擇")
                            sys.exit(0)
                    except Exception as e:
                        if "target frame detached" in "{}".format(e) and reconnection < RECONNECTION_LIMIT:
                            reconnection += 1
                            print(f"\tReconneting ({reconnection}/{RECONNECTION_LIMIT}) ...\n")
                            driver.get(URL)
                            continue
                        else:
                            if "ERR_INTERNET_DISCONNECTED" in "{}".format(e):
                                print(f"Error: 網路連線失敗")
                            elif "target window already closed" in "{}".format(e):
                                print(f"Error: 瀏覽器意外關閉")
                            else:
                                line_number = tb.format_exc().split(",")[1].strip().split()[-1]
                                print(f"\tError: 企鵝物流 - 素材選擇\n\tLine: {line_number}, Error: {e}")
                            sys.exit(0)
                while reconnection <= RECONNECTION_LIMIT:
                    try:
                        # 指定素材頁面
                        WebDriverWait(driver, timeout=2).until(lambda driver: driver.current_url != URL)
                        material_URL = driver.current_url
                        print("\033[F\033[K" * 3 + f"\t正在獲取素材資訊({index_m+1:0>2d}/{len(materials_p):0>2d})：{material_p[0]} ...\n\n")
                        # 選擇數據過濾方式
                        try:
                            Scroll(driver, "up")
                            var = "//div[contains(@class, 'v-expansion-panel') and @aria-expanded]"
                            element = WebDriverWait(driver, timeout=2, poll_frequency=0.1).until(EC.visibility_of_element_located((By.XPATH, var)))
                            if element.get_attribute("aria-expanded") == "false":
                                element.click()
                                try:
                                    Scroll(driver, "up")
                                    var = ".v-input--selection-controls__input"
                                    element = WebDriverWait(driver, timeout=2, poll_frequency=0.1).until(EC.visibility_of_all_elements_located((By.CSS_SELECTOR, var)))[2]
                                    if element.find_element(By.CSS_SELECTOR, "input").get_attribute("aria-checked") == "false":
                                        element.click()
                                except TimeoutException:
                                    print(f"\tTimeout: 僅正在開放")
                                    sys.exit(0)
                                except Exception as e:
                                    if "ERR_INTERNET_DISCONNECTED" in "{}".format(e):
                                        print(f"Error: 網路連線失敗")
                                    elif "target window already closed" in "{}".format(e):
                                        print(f"Error: 瀏覽器意外關閉")
                                    else:
                                        line_number = tb.format_exc().split(",")[1].strip().split()[-1]
                                        print(f"\tError: 僅正在開放\n\tLine: {line_number}, Error: {e}")
                                    sys.exit(0)
                        except TimeoutException:
                            print(f"\tTimeout: 選擇數據過濾方式")
                            sys.exit(0)
                        except Exception as e:
                            if "ERR_INTERNET_DISCONNECTED" in "{}".format(e):
                                print(f"Error: 網路連線失敗")
                            elif "target window already closed" in "{}".format(e):
                                print(f"Error: 瀏覽器意外關閉")
                            else:
                                line_number = tb.format_exc().split(",")[1].strip().split()[-1]
                                print(f"\tError: 選擇數據過濾方式\n\tLine: {line_number}, Error: {e}")
                            sys.exit(0)
                        # 選擇數據排列依據
                        try:
                            Scroll(driver, "up")
                            var = "//th[contains(@aria-label, '单件期望理智')]"
                            element = WebDriverWait(driver, timeout=2, poll_frequency=0.1).until(EC.visibility_of_element_located((By.XPATH, var)))
                            if element.get_attribute("aria-sort") != "ascending":
                                element.click()
                        except TimeoutException:
                            print(f"\tTimeout: 選擇數據排列依據")
                            sys.exit(0)
                        except Exception as e:
                            if "ERR_INTERNET_DISCONNECTED" in "{}".format(e):
                                print(f"Error: 網路連線失敗")
                            elif "target window already closed" in "{}".format(e):
                                print(f"Error: 瀏覽器意外關閉")
                            else:
                                line_number = tb.format_exc().split(",")[1].strip().split()[-1]
                                print(f"\tError: 選擇數據排列依據\n\tLine: {line_number}, Error: {e}")
                            sys.exit(0)
                        # 選擇每頁數量
                        try:
                            Scroll(driver, "down")
                            var = "//div[@class='v-select__selection v-select__selection--comma']"
                            element = WebDriverWait(driver, timeout=2, poll_frequency=0.1).until(EC.visibility_of_element_located((By.XPATH, var)))
                            if element.text != "全部":
                                driver.execute_script("arguments[0].scrollIntoView(true); arguments[0].click();", element)
                                Scroll(driver, "down")
                                var = "//div[@class='v-list-item__title']"
                                elements = WebDriverWait(driver, timeout=2, poll_frequency=0.1).until(EC.visibility_of_all_elements_located((By.XPATH, var)))
                                for element in elements:
                                    if element.text == "全部":
                                        driver.execute_script("arguments[0].scrollIntoView(true); arguments[0].click();", element)
                        except TimeoutException:
                            pass
                        except Exception as e:
                            if "ERR_INTERNET_DISCONNECTED" in "{}".format(e):
                                print(f"Error: 網路連線失敗")
                            elif "target window already closed" in "{}".format(e):
                                print(f"Error: 瀏覽器意外關閉")
                            else:
                                line_number = tb.format_exc().split(",")[1].strip().split()[-1]
                                print(f"\tError: 選擇每頁數量\n\tLine: {line_number}, Error: {e}")
                            sys.exit(0)
                        
                        # 歸零
                        stages_com = []
                        # 儲存關卡名稱
                        var = "span.d-flex.flex-column.ml-2"
                        stages = WebDriverWait(driver, timeout=2, poll_frequency=0.1).until(EC.visibility_of_all_elements_located((By.CSS_SELECTOR, var)))
                        var = "td.px-2.font-weight-bold.monospace.orange--text.text--lighten-1"
                        consumes = WebDriverWait(driver, timeout=2, poll_frequency=0.1).until(EC.visibility_of_all_elements_located((By.CSS_SELECTOR, var)))
                        for index, _ in enumerate(stages):
                            skip = False
                            for str in NoCount:
                                if str in stages[index].text:
                                    skip = True
                                    break
                            if skip:
                                continue
                            stages_com.append(f"{stages[index].text}\n{consumes[index].text}".strip().split("\n"))
                        
                        # 歸零
                        stages_ratio = []
                        # 按關卡
                        for index, stage_com in enumerate(stages_com):
                            temp = reconnection
                            reconnection = 0
                            while reconnection <= RECONNECTION_LIMIT:
                                try:
                                    # 指定素材頁面
                                    WebDriverWait(driver, timeout=2).until(lambda driver: driver.current_url == material_URL)
                                    # 選擇每頁數量
                                    try:
                                        Scroll(driver, "down")
                                        var = "//div[@class='v-select__selection v-select__selection--comma']"
                                        element = WebDriverWait(driver, timeout=2, poll_frequency=0.1).until(EC.visibility_of_element_located((By.XPATH, var)))
                                        if element.text != "全部":
                                            driver.execute_script("arguments[0].scrollIntoView(true); arguments[0].click();", element)
                                            Scroll(driver, "down")
                                            var = "//div[@class='v-list-item__title']"
                                            elements = WebDriverWait(driver, timeout=2, poll_frequency=0.1).until(EC.visibility_of_all_elements_located((By.XPATH, var)))
                                            for element in elements:
                                                if element.text == "全部":
                                                    driver.execute_script("arguments[0].scrollIntoView(true); arguments[0].click();", element)
                                    except TimeoutException:
                                        pass
                                    except Exception as e:
                                        if "ERR_INTERNET_DISCONNECTED" in "{}".format(e):
                                            print(f"Error: 網路連線失敗")
                                        elif "target window already closed" in "{}".format(e):
                                            print(f"Error: 瀏覽器意外關閉")
                                        else:
                                            line_number = tb.format_exc().split(",")[1].strip().split()[-1]
                                            print(f"\tError: 選擇每頁數量\n\tLine: {line_number}, Error: {e}")
                                        sys.exit(0)
                                    # 選擇關卡
                                    var = f"//span[contains(text(), '{stage_com[0]}')]/following-sibling::span[contains(text(), '{stage_com[1]}')]"
                                    driver.execute_script("arguments[0].click();", 
                                        WebDriverWait(driver, timeout=2, poll_frequency=0.1).until(EC.visibility_of_element_located((By.XPATH, var))))
                                    reconnection = temp
                                    break
                                except TimeoutException as e:
                                    if reconnection < RECONNECTION_LIMIT:
                                        reconnection += 1
                                        print(f"\tReconneting ({reconnection}/{RECONNECTION_LIMIT}) ...\n")
                                        driver.get(material_URL)
                                        continue
                                    else:
                                        print(f"\tTimeout: {material_p[0]}")
                                        sys.exit(0)
                                except Exception as e:
                                    if "target frame detached" in "{}".format(e) and reconnection < RECONNECTION_LIMIT:
                                        reconnection += 1
                                        print(f"\tReconneting ({reconnection}/{RECONNECTION_LIMIT}) ...\n")
                                        driver.get(material_URL)
                                        continue
                                    else:
                                        if "ERR_INTERNET_DISCONNECTED" in "{}".format(e):
                                            print(f"Error: 網路連線失敗")
                                        elif "target window already closed" in "{}".format(e):
                                            print(f"Error: 瀏覽器意外關閉")
                                        else:
                                            line_number = tb.format_exc().split(",")[1].strip().split()[-1]
                                            print(f"\tError: {material_p[0]}\n\tLine: {line_number}, Error: {e}")
                                        sys.exit(0)
                            temp = reconnection
                            reconnection = 0
                            while reconnection <= RECONNECTION_LIMIT:
                                try:
                                    # 指定關卡頁面
                                    WebDriverWait(driver, timeout=2).until(lambda driver: driver.current_url != material_URL)
                                    print("\033[F\033[K" * 2 + f"\t\t正在獲取關卡資訊({index+1:0>2d}/{len(stages_com):0>2d})：{stage_com[0]} {stage_com[1]} ...\n")
                                    # 選擇每頁數量
                                    try:
                                        Scroll(driver, "down")
                                        var = "//div[@class='v-select__selection v-select__selection--comma']"
                                        element = WebDriverWait(driver, timeout=2, poll_frequency=0.1).until(EC.visibility_of_all_elements_located((By.XPATH, var)))[0]
                                        if element.text != "全部":
                                            driver.execute_script("arguments[0].scrollIntoView(true); arguments[0].click();", element)
                                            Scroll(driver, "down")
                                            var = "//div[@class='v-list-item__title']"
                                            elements = WebDriverWait(driver, timeout=2, poll_frequency=0.1).until(EC.visibility_of_all_elements_located((By.XPATH, var)))
                                            for element in elements:
                                                if element.text == "全部":
                                                    driver.execute_script("arguments[0].scrollIntoView(true); arguments[0].click();", element)
                                    except TimeoutException:
                                        pass
                                    except Exception as e:
                                        if "ERR_INTERNET_DISCONNECTED" in "{}".format(e):
                                            print(f"Error: 網路連線失敗")
                                        elif "target window already closed" in "{}".format(e):
                                            print(f"Error: 瀏覽器意外關閉")
                                        else:
                                            line_number = tb.format_exc().split(",")[1].strip().split()[-1]
                                            print(f"\t\tError: 選擇每頁數量\n\t\tLine: {line_number}, Error: {e}")
                                        sys.exit(0)
                                    # 儲存各素材掉落率
                                    var1 = "//tr[@class='']//span[contains(@class, 'item-name--text')]"
                                    var2 = "//tr[@class='']//td[contains(@class, 'd-flex align-center justify-start fill-height px-2 font-weight-bold monospace')]//span"
                                    names = driver.find_elements(By.XPATH, var1)
                                    percents = driver.find_elements(By.XPATH, var2)
                                    # 歸零
                                    datas = []
                                    earn = 0
                                    for num in range(len(names)):
                                        if names[num].text != "家具" and "给" not in names[num].text:
                                            print("\033[F\033[K" + f"\t\t\t正在獲取掉落：{names[num].text} ...")
                                            datas.append(f"{names[num].text} {percents[num].text}".split())
                                    print("\033[F\033[K" + f"\t\t\t正在計算關卡理智比 ...")
                                    for data in datas:
                                        for material_all in materials_all:
                                            if material_all[0] == data[0]:
                                                earn += float(material_all[2]) * float(data[1][:-1]) / 100
                                    stages_ratio.append([material_p[1], stage_com[0], stage_com[1], round(earn/float(stage_com[2]), 10)])
                                    # 返回指定素材
                                    driver.back()
                                    reconnection = temp
                                    break
                                except TimeoutException as e:
                                    if reconnection < RECONNECTION_LIMIT:
                                        reconnection += 1
                                        print(f"\t\tReconneting ({reconnection}/{RECONNECTION_LIMIT}) ...\n")
                                        driver.get(material_URL)
                                        continue
                                    else:
                                        print(f"\t\tTimeout: {material_p[0]}：{stage_com[0]} {stage_com[1]}")
                                        sys.exit(0)
                                except Exception as e:
                                    if "target frame detached" in "{}".format(e) and reconnection < RECONNECTION_LIMIT:
                                        reconnection += 1
                                        print(f"\t\tReconneting ({reconnection}/{RECONNECTION_LIMIT}) ...\n")
                                        driver.get(material_URL)
                                        continue
                                    else:
                                        if "ERR_INTERNET_DISCONNECTED" in "{}".format(e):
                                            print(f"Error: 網路連線失敗")
                                        elif "target window already closed" in "{}".format(e):
                                            print(f"Error: 瀏覽器意外關閉")
                                        else:
                                            line_number = tb.format_exc().split(",")[1].strip().split()[-1]
                                            print(f"\t\tError: {material_p[0]}：{stage_com[0]} {stage_com[1]}\n\t\tLine: {line_number}, Error: {e}")
                                        sys.exit(0)
                        # 歸零
                        material_ratio = ["", "", "", 0]
                        for stage_ratio in stages_ratio:
                            if stage_ratio[3] > material_ratio[3]:
                                material_ratio = stage_ratio[:]
                        materials_ratio.append(material_ratio)
                        # 返回所有素材
                        driver.get(URL)
                        reconnection = 0
                        break
                    except TimeoutException as e:
                        if reconnection < RECONNECTION_LIMIT:
                            reconnection += 1
                            print(f"\tReconneting ({reconnection}/{RECONNECTION_LIMIT}) ...\n")
                            driver.get(material_URL)
                            continue
                        else:
                            print(f"\tTimeout: {material_p[0]}")
                            sys.exit(0)
                    except Exception as e:
                        if "target frame detached" in "{}".format(e) and reconnection < RECONNECTION_LIMIT:
                            reconnection += 1
                            print(f"\tReconneting ({reconnection}/{RECONNECTION_LIMIT}) ...\n")
                            driver.get(material_URL)
                            continue
                        else:
                            if "ERR_INTERNET_DISCONNECTED" in "{}".format(e):
                                print(f"Error: 網路連線失敗")
                            elif "target window already closed" in "{}".format(e):
                                print(f"Error: 瀏覽器意外關閉")
                            else:
                                line_number = tb.format_exc().split(",")[1].strip().split()[-1]
                                print(f"\tError: {material_p[0]}\n\tLine: {line_number}, Error: {e}")
                            sys.exit(0)
            
            # 綜合素材最高效率關卡更新
            print("\033[F\033[K" * 4 + "\"綜合素材最高效率關卡\" 更新中 ...")
            try:
                app = xw.App(visible=False)
                workbook = app.books.open(FILE_PATH)
                workbook.app.calculate()
                workbook.save()
                workbook.close()
                app.quit()
                workbook = opxl.load_workbook(FILE_PATH, data_only = False)
                sheet = "打素材"
                worksheet = workbook[sheet]
                for index, row in enumerate(worksheet.iter_rows(min_col=1, min_row=2), start = 2):
                    for data in materials_ratio:
                        if data[0][1:] == row[0].value:
                            worksheet.cell(row=index, column=row[0].column+4, value=float(data[3]))
                            if "标准" in data[1]:
                                worksheet.cell(row=index, column=row[0].column+3, value=f"{data[2]}(標準)")
                            elif "磨难" in data[1]:
                                worksheet.cell(row=index, column=row[0].column+3, value=f"{data[2]}(磨難)")
                            else:
                                worksheet.cell(row=index, column=row[0].column+3, value=data[2])
                            break
                workbook.save(FILE_PATH)
                workbook.close()
            except Exception as e:
                if "Permission denied" in "{}".format(e):
                    print(f"Error: \"綜合素材最高效率關卡\" 更新失敗，請先關閉檔案 \"{FILE_PATH}\"")
                else:
                    line_number = tb.format_exc().split(",")[1].strip().split()[-1]
                    print(f"Error: \"綜合素材最高效率關卡\" 更新\nLine: {line_number}, Error: {e}")
                sys.exit(0)
            print("\033[F\033[K\"綜合素材最高效率關卡\" 更新完成")
    except KeyboardInterrupt:
        print("Keyboard interruptted!")
        sys.exit(0)
    except TimeoutException:
        print(f"Timeout: 企鵝物流 - 素材選擇")
        sys.exit(0)
    except Exception as e:
        if "ERR_INTERNET_DISCONNECTED" in "{}".format(e):
            print(f"Error: 網路連線失敗")
        elif "target window already closed" in "{}".format(e):
            print(f"Error: 瀏覽器意外關閉")
        else:
            line_number = tb.format_exc().split(",")[1].strip().split()[-1]
            print(f"Error: 企鵝物流 - 素材選擇\nLine: {line_number}, Error: {e}")
        sys.exit(0)
    finally:
        if driver:
            print("Closing connection ...")
            driver.quit()
        print("\033[F\033[KConnection closed!")
    
    # 打印結束執行時間
    print(f"\nEnd time  : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

